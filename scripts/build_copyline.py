# -*- coding: utf-8 -*-
"""
CopyLine adapter — сборщик по шаблону CS (использует scripts/cs/core.py).
Задача адаптера: забрать данные поставщика (sitemap + сайт) и отдать в CS ядро список OfferOut.
"""

from __future__ import annotations

import io
import os
import re
import time
import random
import hashlib
from datetime import datetime, timedelta

# Логи (можно выключить: VERBOSE=0)
def _pick_copyline_best_picture(pictures: list[str]) -> list[str]:
    """CopyLine: оставить только реальные фото товара.
    Берём только картинки из img_products, чистим мусор, сохраняем порядок (full_* сначала).
    """
    if not pictures:
        return [PLACEHOLDER_PIC]

    cleaned: list[str] = []
    seen: set[str] = set()

    for p in pictures:
        if not p:
            continue
        p = str(p).strip()
        if not p:
            continue

        # только реальные фото товаров (без логотипов/иконок/печатных и т.п.)
        if "/components/com_jshopping/files/img_products/" not in p.replace("\\", "/"):
            continue

        # нормализуем HTML-экранку
        p = p.replace("&amp;", "&")

        if p in seen:
            continue
        seen.add(p)
        cleaned.append(p)

    if not cleaned:
        return [PLACEHOLDER_PIC]

    def is_full(u: str) -> bool:
        base = u.split("/")[-1]
        return base.startswith("full_") or "/full_" in u

    fulls = [u for u in cleaned if is_full(u)]
    normals = [u for u in cleaned if not is_full(u)]

    # если на странице есть только обычное фото — его и оставим
    return (fulls + normals) if (fulls + normals) else [PLACEHOLDER_PIC]
def _pick_copyline_picture(pics: list[str]) -> list[str]:
    """# CopyLine: одна картинка на товар — full_ если есть, иначе обычная. Только img_products."""
    if not pics:
        return []

    def norm(u: str) -> str:
        u = (u or "").strip()
        u = u.split("#", 1)[0]
        return u

    candidates: list[str] = []
    for u in pics:
        u = norm(u)
        if not u:
            continue
        if "components/com_jshopping/files/img_products/" not in u:
            continue
        if "/img_products/thumb_" in u:
            continue
        candidates.append(u)

    if not candidates:
        return []

    # full_ приоритет
    for u in candidates:
        base = u.rsplit("/", 1)[-1]
        if base.startswith("full_"):
            return [u]

    return [candidates[0]]

VERBOSE = os.environ.get("VERBOSE", "0") in ("1","true","True","yes","YES")

def log(*args, **kwargs) -> None:
    # Печать логов (в Actions удобно оставлять краткие метки)
    # Поддерживаем kwargs типа flush/end/sep, чтобы не ловить TypeError.
    if VERBOSE:
        if "flush" not in kwargs:
            kwargs["flush"] = True
        print(*args, **kwargs)

import requests
from bs4 import BeautifulSoup
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None  # sitemap-режим работает без openpyxl

from cs.core import (
    CURRENCY_ID_DEFAULT,
    OfferOut,
    compute_price,
    ensure_footer_spacing,
    make_feed_meta,
    make_footer,
    make_header,
    now_almaty,
    validate_cs_yml,
    write_if_changed,
)

# -----------------------------
# Настройки
# -----------------------------
SUPPLIER_NAME = "CopyLine"
SUPPLIER_URL_DEFAULT = "https://copyline.kz/goods.html"
BASE_URL = "https://copyline.kz"

SITEMAP_URL_DEFAULT = f"{BASE_URL}/site-map.html?id=1&view=html"
SITEMAP_URL = os.getenv("SITEMAP_URL", SITEMAP_URL_DEFAULT)

XLSX_URL = os.getenv("XLSX_URL", f"{BASE_URL}/files/price-CLA.xlsx")

# Вариант C: фильтрация CopyLine по префиксам названия (строго с начала строки)
# Важно для стабильного ассортимента и чтобы не тянуть UPS/прочее из прайса.
COPYLINE_INCLUDE_PREFIXES = ['Drum', 'Девелопер', 'Драм-картридж', 'Драм-юниты', 'Кабель сетевой', 'Картридж', 'Картриджи', 'Термоблок', 'Тонер-картридж', 'Чернила']



OUT_FILE = os.getenv("OUT_FILE", "docs/copyline.yml")
OUTPUT_ENCODING = (os.getenv("OUTPUT_ENCODING", "utf-8") or "utf-8").strip() or "utf-8"
NO_CRAWL = (os.getenv("NO_CRAWL", "0") or "0").strip().lower() in ("1", "true", "yes", "y", "on")
MAX_CATEGORY_PAGES = int(os.getenv("MAX_CATEGORY_PAGES", "25") or "25")  # лимит страниц на категорию
MAX_CRAWL_MINUTES = int(os.getenv("MAX_CRAWL_MINUTES", "60") or "60")    # общий лимит времени обхода сайта
# Регулярка для карточек товара (не категорий)
PRODUCT_RE = re.compile(r"/goods/[^/]+\.html(?:[?#].*)?$", flags=re.I)

# Параллелизм обхода сайта
MAX_WORKERS = int(os.getenv("MAX_WORKERS", "6") or "6")




VENDORCODE_PREFIX = (os.getenv("VENDORCODE_PREFIX") or "CL").strip()
PUBLIC_VENDOR = (os.getenv("PUBLIC_VENDOR") or SUPPLIER_NAME).strip() or SUPPLIER_NAME

HTTP_TIMEOUT = float(os.getenv("HTTP_TIMEOUT", "30"))
REQUEST_DELAY_MS = int(os.getenv("REQUEST_DELAY_MS", "60"))
# HTTP headers (нужно для requests.get; иначе некоторые ответы могут быть урезаны)
UA = {
    "User-Agent": os.getenv(
        "HTTP_UA",
        "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    ),
    "Accept": "*/*",
    "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.7,en;q=0.5",
    "Connection": "keep-alive",
}



def _sleep_jitter(ms: int) -> None:
    d = max(0.0, ms / 1000.0)
    time.sleep(d * (1.0 + random.uniform(-0.15, 0.15)))


def http_get(url: str, tries: int = 3, min_bytes: int = 0) -> Optional[bytes]:
    delay = max(0.1, REQUEST_DELAY_MS / 1000.0)
    last = None
    for _ in range(max(1, tries)):
        try:
            r = requests.get(url, headers=UA, timeout=HTTP_TIMEOUT)
            if r.status_code == 200 and (len(r.content) >= min_bytes):
                return r.content
            last = f"http {r.status_code} size={len(r.content)}"
        except Exception as e:
            last = repr(e)
        _sleep_jitter(int(delay * 1000))
        delay *= 1.6
    log(f"[http] fail: {url} | {last}")
    return None



def http_post(url: str, data: Dict[str, Any], tries: int = 3, min_bytes: int = 0) -> Optional[bytes]:
    delay = max(0.1, REQUEST_DELAY_MS / 1000.0)
    last = None
    for _ in range(max(1, tries)):
        try:
            r = requests.post(url, data=data, headers=UA, timeout=HTTP_TIMEOUT)
            if r.status_code == 200 and (len(r.content) >= min_bytes):
                return r.content
            last = f"http {r.status_code} size={len(r.content)}"
        except Exception as e:
            last = repr(e)
        _sleep_jitter(int(delay * 1000))
        delay *= 1.6
    log(f"[http] POST fail: {url} | {last}")
    return None

def soup_of(b: bytes) -> BeautifulSoup:
    return BeautifulSoup(b, "html.parser")



def _pick_best_copyline_pic(urls: List[str]) -> Optional[str]:
    """CopyLine: оставляем только реальные фото товара (img_products).
    full_ приоритет, затем обычная, thumb_ в конце. Возвращаем 1 URL."""
    cleaned: List[str] = []
    seen = set()
    for u in urls or []:
        nu = u
        if not nu:
            continue
        if "/components/com_jshopping/files/img_products/" not in nu:
            continue
        if nu in seen:
            continue
        seen.add(nu)
        cleaned.append(nu)
    if not cleaned:
        return None

    def score(u: str) -> int:
        fname = u.rsplit("/", 1)[-1]
        if fname.startswith("full_") or "/img_products/full_" in u:
            return 30
        if fname.startswith("thumb_") or "/img_products/thumb_" in u:
            return 10
        return 20

    cleaned.sort(key=score, reverse=True)
    return cleaned[0]


def _parse_search_results(html_bytes: bytes) -> List[Dict[str, Any]]:
    s = soup_of(html_bytes)
    results: List[Dict[str, Any]] = []
    seen_url = set()

    for a in s.find_all("a", href=True):
        href = safe_str(a.get("href")).strip()
        if not href:
            continue
        if "/goods/" not in href or ".html" not in href:
            continue

        if href.startswith("/"):
            url = BASE_URL + href
        elif href.startswith("//"):
            url = "https:" + href
        elif href.startswith("http"):
            url = href
        else:
            url = BASE_URL.rstrip("/") + "/" + href.lstrip("/")

        if url in seen_url:
            continue
        seen_url.add(url)

        title = safe_str(a.get_text(" ", strip=True))
        imgs: List[str] = []
        for img in a.find_all("img"):
            for attr in ("data-src", "data-original", "data-lazy", "src", "srcset"):
                v = img.get(attr)
                if v:
                    imgs.append(safe_str(v))
                    break

        if not imgs and a.parent:
            for img in a.parent.find_all("img"):
                for attr in ("data-src", "data-original", "data-lazy", "src", "srcset"):
                    v = img.get(attr)
                    if v:
                        imgs.append(safe_str(v))
                        break

        results.append({"url": url, "title": title, "imgs": imgs})

    return results


def _choose_best_search_hit(hits: List[Dict[str, Any]], sku: str) -> Optional[Dict[str, Any]]:
    if not hits:
        return None
    key = norm_ascii(sku)

    def hit_score(h: Dict[str, Any]) -> int:
        u = norm_ascii(safe_str(h.get("url")))
        t = norm_ascii(safe_str(h.get("title")))
        sc = 0
        if key and key in u:
            sc += 100
        if key and key in t:
            sc += 80
        if len(hits) == 1:
            sc += 5
        return sc

    return max(hits, key=hit_score)


def _search_copyline_for_sku(sku: str) -> Optional[Dict[str, Any]]:
    """Ищем товар по артикулу (колонка B в XLSX) через официальный поиск CopyLine.
    Сначала пытаемся взять фото прямо из выдачи, если не получилось — открываем карточку товара."""
    b = http_post("https://copyline.kz/product-search/result.html", data={"search": sku}, tries=3)
    if not b:
        return None

    hits = _parse_search_results(b)
    best = _choose_best_search_hit(hits, sku)
    if not best:
        return None

    # фото из выдачи
    pic = _pick_best_copyline_pic(best.get("imgs") or [])
    out: Dict[str, Any] = {
        "sku": sku,
        "title": safe_str(best.get("title") or ""),
        "desc": "",
        "pic": pic or "",
        "pics": [pic] if pic else [],
        "params": [],
        "url": safe_str(best.get("url") or ""),
    }

    # если фото не нашли — лезем в карточку
    if (not pic) and out["url"]:
        p = parse_product_page(out["url"])
        if p:
            p["sku"] = sku  # принудительно из XLSX
            best_pic = _pick_best_copyline_pic(list(p.get("pics") or []) + ([safe_str(p.get("pic"))] if p.get("pic") else []))
            if best_pic:
                p["pic"] = best_pic
                p["pics"] = [best_pic]
            return p

    return out

def norm_ascii(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (s or "").lower())


def _norm_sku_variants(raw: str) -> set[str]:
    """# CopyLine: нормализация артикула/sku для сравнения (детерминированно)"""
    r = (raw or "").strip()
    if not r:
        return set()
    r = r.replace(" ", "")
    variants = {r, r.replace("-", ""), r.replace("_", "")}
    r0 = r.lstrip("0")
    if r0 and r0 != r:
        variants.add(r0)
        variants.add(r0.replace("-", ""))
        variants.add(r0.replace("_", ""))
    if re.fullmatch(r"[Cc]\d+", r):
        variants.add(r[1:])
    if re.fullmatch(r"\d+", r):
        variants.add("C" + r)
    return {norm_ascii(v) for v in variants if v}

def _sku_matches(raw_v: str, page_sku: str) -> bool:
    """# CopyLine: страница товара подходит только если sku совпадает с vendorCode_raw"""
    want = _norm_sku_variants(raw_v)
    have = _norm_sku_variants(page_sku)
    return bool(want and have and (want & have))

def parse_price_tenge(text: str) -> int:
    """Парсинг цены в тг из текста: '7 051 тг.' -> 7051."""
    if not text:
        return 0
    s = str(text)
    m = re.search(r"(\d[\d\s]{1,15})\s*(?:тг|тенге|₸)", s, flags=re.I)
    if not m:
        return 0
    num = re.sub(r"\s+", "", m.group(1))
    try:
        return int(num)
    except Exception:
        return 0


def parse_price_digits(text: str) -> int:
    """Парсинг числа из блока цены.
    Не привязываемся к 'тг', чтобы не ломаться на кривой кодировке.
    Поддерживаем варианты '7 051', '7051', '7051.00'.
    """
    if not text:
        return 0
    s = str(text)
    # если есть десятичная часть, берём целую
    m = re.search(r"(\d[\d\s]{0,15})(?:[\.,]\d{1,2})", s)
    if m:
        num = re.sub(r"\s+", "", m.group(1))
        try:
            return int(num)
        except Exception:
            return 0
    num = re.sub(r"[^0-9]+", "", s)
    if not num:
        return 0
    try:
        return int(num)
    except Exception:
        return 0
def title_clean(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"\s*\((?:Артикул|SKU|Код)\s*[:#]?\s*[^)]+\)\s*$", "", s, flags=re.I)
    return re.sub(r"\s{2,}", " ", s).strip()[:200]


def safe_str(x: Any) -> str:
    return (str(x).strip() if x is not None else "")


def to_number(x: Any) -> Optional[float]:
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if not s:
        return None
    s = s.replace("\xa0", " ").replace(" ", "")
    s = s.replace(",", ".")
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    if not m:
        return None
    try:
        return float(m.group(0))
    except Exception:
        return None


def parse_stock_to_bool(x: Any) -> bool:
    if x is None:
        return False
    if isinstance(x, (int, float)):
        return float(x) > 0
    s = str(x).strip()
    if not s:
        return False
    s_low = s.lower()
    if s_low in ("-", "нет", "0", "0.0"):
        return False
    # "<10", ">5", "есть", "1-2" — считаем как наличие
    if re.search(r"\d", s_low):
        return True
    if "есть" in s_low:
        return True
    return False


def oid_from_vendor_code_raw(raw: str) -> str:
    raw = (raw or "").strip()
    raw = raw.replace("–", "-").replace("/", "-").replace("\\", "-")
    raw = re.sub(r"\s+", "", raw)
    raw = re.sub(r"[^A-Za-z0-9_.-]+", "", raw)
    raw = raw.strip("-.")
    if not raw:
        # аварийный вариант (стабильный, но без исходного кода)
        h = hashlib.md5((raw or "empty").encode("utf-8", errors="ignore")).hexdigest()[:10].upper()
        return f"{VENDORCODE_PREFIX}{h}"
    return f"{VENDORCODE_PREFIX}{raw}"


def compile_startswith_patterns(kws: Sequence[str]) -> List[re.Pattern]:
    # строго с начала строки, чтобы не тянуть мусорные позиции
    out: List[re.Pattern] = []
    for kw in kws:
        kw = kw.strip()
        if not kw:
            continue
        out.append(re.compile(r"^\s*" + re.escape(kw).replace(r"\ ", " ") + r"(?!\w)", re.I))
    return out


def title_startswith_strict(title: str, patterns: Sequence[re.Pattern]) -> bool:
    return bool(title) and any(p.search(title) for p in patterns)


def _is_allowed_prefix(name: str) -> bool:
    """# CopyLine: фильтр по первому слову/префиксу из COPYLINE_INCLUDE_PREFIXES."""
    n = (name or '').strip()
    if not n:
        return False
    if not COPYLINE_INCLUDE_PREFIXES:
        return True
    low = n.lower()
    for p in COPYLINE_INCLUDE_PREFIXES:
        pl = (p or '').strip().lower()
        if not pl:
            continue
        if low == pl or low.startswith(pl + ' ') or low.startswith(pl + '-'):
            return True
    return False

def detect_header_two_row(rows: List[List[Any]], scan_rows: int = 60) -> Tuple[int, int, Dict[str, int]]:
    def low(x: Any) -> str:
        return safe_str(x).lower()

    for i in range(min(scan_rows, len(rows) - 1)):
        row0 = [low(c) for c in rows[i]]
        row1 = [low(c) for c in rows[i + 1]]

        if any("номенклатура" in c for c in row0):
            name_col = next((j for j, c in enumerate(row0) if "номенклатура" in c), None)
            vendor_col = next((j for j, c in enumerate(row1) if "артикул" in c), None)
            price_col = next((j for j, c in enumerate(row1) if "цена" in c or "опт" in c), None)
            unit_col = next((j for j, c in enumerate(row1) if c.strip().startswith("ед")), None)
            stock_col = (
                next((j for j, c in enumerate(row0) if "остаток" in c), None)
                or next((j for j, c in enumerate(row1) if "остаток" in c), None)
            )
            if name_col is not None and vendor_col is not None and price_col is not None:
                idx = {"name": name_col, "vendor_code": vendor_col, "price": price_col}
                if stock_col is not None:
                    idx["stock"] = stock_col
                return i, i + 1, idx

    return -1, -1, {}



def _derive_kind(title: str) -> str:
    t = (title or "").strip().lower()
    if not t:
        return ""
    if t.startswith("тонер-картридж") or t.startswith("тонер картридж"):
        return "Тонер-картридж"
    if t.startswith("картридж"):
        return "Картридж"
    if t.startswith("кабель сетевой"):
        return "Кабель сетевой"
    if t.startswith("термоблок"):
        return "Термоблок"
    if t.startswith("термоэлемент"):
        return "Термоэлемент"
    if t.startswith("девелопер") or t.startswith("developer"):
        return "Девелопер"
    if t.startswith("драм") or t.startswith("drum"):
        return "Драм-картридж"
    return ""

def _merge_params(existing: List[Tuple[str, str]], add: List[Tuple[str, str]]) -> List[Tuple[str, str]]:
    # Склеиваем параметры без дублей, и выкидываем мусорные ключи (например, "3").
    seen = set()
    out: List[Tuple[str, str]] = []

    def push(k: str, v: str) -> None:
        kk = (k or "").strip()
        vv = (v or "").strip()
        if not kk or not vv:
            return
        if kk.isdigit():
            return
        key = (kk.lower(), vv.lower())
        if key in seen:
            return
        seen.add(key)
        out.append((kk, vv))

    for k, v in (existing or []):
        push(k, v)
    for k, v in (add or []):
        push(k, v)

    return out


def parse_xlsx_items(xlsx_bytes: bytes) -> Tuple[int, List[Dict[str, Any]]]:
    if load_workbook is None:

        raise RuntimeError("CopyLine: openpyxl не установлен (XLSX-режим недоступен). Используй sitemap-режим.")

    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    sheet = max(wb.sheetnames, key=lambda n: wb[n].max_row * max(1, wb[n].max_column))
    ws = wb[sheet]
    rows = [[c for c in r] for r in ws.iter_rows(values_only=True)]
    log(f"[xls] sheet={sheet} rows={len(rows)}")

    row0, row1, idx = detect_header_two_row(rows)
    if row0 < 0:
        raise RuntimeError("Не удалось распознать шапку в XLSX.")

    data_start = row1 + 1
    name_col, vendor_col, price_col = idx["name"], idx["vendor_code"], idx["price"]
    stock_col = idx.get("stock")
    unit_col = idx.get("unit")
    kws = COPYLINE_INCLUDE_PREFIXES
    start_patterns = compile_startswith_patterns(kws)
    source_rows = sum(1 for r in rows[data_start:] if any(v is not None and str(v).strip() for v in r))

    out: List[Dict[str, Any]] = []
    for r in rows[data_start:]:
        name_raw = r[name_col]
        if not name_raw:
            continue
        title = title_clean(safe_str(name_raw))
        if not title_startswith_strict(title, start_patterns):
            continue

        dealer = to_number(r[price_col])
        if dealer is None or dealer <= 0:
            continue

        v_raw = r[vendor_col]
        vcode = safe_str(v_raw)
        if not vcode:
            # иногда артикул спрятан в названии
            m = re.search(r"[A-ZА-Я0-9]{2,}(?:[-/–][A-ZА-Я0-9]{2,})?", title.upper())
            if m:
                vcode = m.group(0).replace("–", "-").replace("/", "-")
        if not vcode:
            continue

        available = True
        if stock_col is not None and stock_col < len(r):
            available = parse_stock_to_bool(r[stock_col])

        out.append(
            {
                "title": title,
                "vendorCode_raw": vcode,
                "dealer_price": int(round(float(dealer))),
                "available": bool(available),
                "stock_raw": safe_str(r[stock_col]).strip() if (stock_col is not None and stock_col < len(r)) else "",
                "unit_raw": safe_str(r[unit_col]).strip() if (unit_col is not None and unit_col < len(r)) else "",
            }
        )

    log(f"[xls] source_rows={source_rows} filtered={len(out)}")
    return source_rows, out


# -----------------------------
# Сайт: индексация карточек (картинки + описание + характеристики)
# -----------------------------
def normalize_url(url: Optional[str]) -> Optional[str]:
    """Нормализация URL картинки (CopyLine).
    Важно: не превращаем автоматически в full_ (иначе получаем несуществующие ссылки).
    Делаем только: absolute + убираем #/? + режем srcset хвост."""
    if not url:
        return None
    u = safe_str(url).strip()
    if not u:
        return None
    # srcset может прийти как "url 2x"
    if " " in u and (u.startswith("http") or u.startswith("/") or u.startswith("//")):
        u = u.split(" ", 1)[0].strip()
    if "#" in u:
        u = u.split("#", 1)[0]
    if "?" in u:
        u = u.split("?", 1)[0]
    if not u:
        return None
    if u.startswith("//"):
        u = "https:" + u
    elif u.startswith("/"):
        u = BASE_URL + u
    elif u.startswith("http://"):
        u = "https://" + u[len("http://"):]
    return u

def extract_kv_pairs_from_text(text: str) -> List[Tuple[str, str]]:
    # очень мягкий парсер "Ключ: значение" в тексте
    out: List[Tuple[str, str]] = []
    for ln in (text or "").splitlines():
        ln = ln.strip().strip("•-–—")
        if not ln:
            continue
        if ":" in ln:
            k, v = ln.split(":", 1)
            k = k.strip()
            v = v.strip()
            if k and v and len(k) <= 80 and len(v) <= 240:
                out.append((k, v))
    return out


def _copyline_full_only(pics: list[str]) -> list[str]:
    """# CopyLine: если есть full_ — оставляем только full_ (обычные дублируют)."""
    if not pics:
        return []
    full: list[str] = []
    other: list[str] = []
    seen = set()
    for u in pics:
        u = (u or "").strip()
        if not u:
            continue
        if u in seen:
            continue
        seen.add(u)
        base = u.rsplit("/", 1)[-1]
        if base.startswith("full_"):
            full.append(u)
        else:
            other.append(u)
    return full if full else other

def parse_product_page(url: str) -> Optional[Dict[str, Any]]:
    b = http_get(url, tries=3)
    if not b:
        return None
    s = soup_of(b)

    # SKU
    sku = ""
    skuel = s.find(attrs={"itemprop": "sku"})
    if skuel:
        sku = safe_str(skuel.get_text(" ", strip=True))

    # jshopping: Артикул часто лежит тут: <span id="product_code">101942</span>
    if not sku:
        pc = s.find(id="product_code")
        if pc:
            sku = safe_str(pc.get_text(" ", strip=True))
    if not sku:
        txt = s.get_text(" ", strip=True)
        m = re.search(r"(?:Артикул|SKU|Код товара|Код)\s*[:#]?\s*([A-Za-z0-9\-\._/]{2,})", txt, flags=re.I)
        if m:
            sku = m.group(1)


    if not sku:
        return None

    # Title
    h = s.find(["h1", "h2"], attrs={"itemprop": "name"}) or s.find("h1") or s.find("h2")
    title = title_clean(safe_str(h.get_text(" ", strip=True) if h else ""))
    # Picture (на сайте почти всегда есть фото; вытаскиваем максимально надёжно)

    cand: list[str] = []


    # 0) основная картинка (как на странице): <a class="lightbox" id="main_image_full_..."> href="...full_*.jpg"

    a_full = s.select_one('a.lightbox[id^="main_image_full_"]')

    if a_full and a_full.get("href"):

        cand.append(safe_str(a_full["href"]))


    # 1) og:image (обычно ведёт на img_products/*.jpg)

    ogi = s.find("meta", attrs={"property": "og:image"})

    if ogi and ogi.get("content"):

        cand.append(safe_str(ogi["content"]))


    # 2) rel=image_src

    lnk = s.find("link", attrs={"rel": "image_src"})

    if lnk and lnk.get("href"):

        cand.append(safe_str(lnk["href"]))


    # 3) main_image_* / itemprop=image

    img_main = s.select_one('img[id^="main_image_"]') or s.find("img", attrs={"itemprop": "image"})

    if img_main:

        for a in ("data-src", "data-original", "data-lazy", "src", "srcset"):

            v = img_main.get(a)

            if v:

                cand.append(safe_str(v))

                break


    # 4) любые img на странице (отбираем только похожие на фото товара)

    for img in s.find_all("img"):

        for a in ("data-src", "data-original", "data-lazy", "src", "srcset"):

            t = safe_str(img.get(a))

            if not t:

                continue

            if "thumb_" in t:

                continue

            if any(k in t for k in ("img_products", "jshopping", "/products/", "/img/")) or re.search(r"\.(?:jpg|jpeg|png|webp)(?:\?|$)", t, flags=re.I):

                cand.append(t)

                break


    # 5) иногда большая картинка лежит в <a href="...full_...jpg">

    for a in s.find_all("a"):

        href = safe_str(a.get("href"))

        if not href:

            continue

        if "thumb_" in href:

            continue

        if ("img_products" in href) or re.search(r"\.(?:jpg|jpeg|png|webp)(?:\?|$)", href, flags=re.I):

            cand.append(href)
    # Соберём все картинки без подмены на full_ (иначе легко получить 404)
    pics_raw: List[str] = []
    for t in cand:
        t = (t or "").strip()
        if not t or t.startswith("data:"):
            continue
        pics_raw.append(t)

    pics: List[str] = []
    seen: Set[str] = set()
    for t in pics_raw:
        u = t
        if not u or u.startswith("data:"):
            continue
        if u in seen:
            continue
        seen.add(u)
        pics.append(u)

    # все фото (только img_products, full_* приоритет)
    pics = _pick_copyline_best_picture(pics)
    pic = (pics[0] if pics else PLACEHOLDER_PIC)



    # Description + params
    desc_txt = ""
    params: List[Tuple[str, str]] = []

    block = (
        s.select_one('div[itemprop="description"].jshop_prod_description')
        or s.select_one("div.jshop_prod_description")
        or s.select_one('[itemprop="description"]')
    )
    if block:
        desc_txt = block.get_text("\n", strip=True)
        params.extend(extract_kv_pairs_from_text(desc_txt))

    # Table specs (если есть)
    table = s.find("table")
    if table:
        for tr in table.find_all("tr"):
            tds = tr.find_all(["td", "th"])
            if len(tds) >= 2:
                k = safe_str(tds[0].get_text(" ", strip=True))
                v = safe_str(tds[1].get_text(" ", strip=True))
                if k and v and len(k) <= 80 and len(v) <= 240:
                    params.append((k, v))

    # Удалим дубли
    seen = set()
    params2: List[Tuple[str, str]] = []
    for k, v in params:
        kk = k.strip()
        vv = v.strip()
        if not kk or not vv:
            continue
        key = (kk.lower(), vv.lower())
        if key in seen:
            continue
        seen.add(key)
        params2.append((kk, vv))

    # Price (тг)
    price_raw = 0
    price_src = ""

    # Важно: цену берём ТОЛЬКО из основного блока товара (.productfull),
    # иначе схватим цену из "похожие товары"/"рекомендуемые" (там .jshop_price).
    main = s.select_one(".productfull") or s

    # 1) Основная цена на CopyLine обычно в <span id="block_price">...</span>
    pr = main.select_one("#block_price") or main.find(id="block_price")
    if pr:
        price_raw = parse_price_digits(pr.get_text(" ", strip=True))
        if price_raw:
            price_src = "block_price"

    # 2) Если структура другая — пробуем блок цены
    if not price_raw:
        box = main.select_one(".prod_price") or main.select_one(".jshop_prod_price")
        if box:
            # текущую цену часто кладут в span с id/классом, старую — в old_price
            cand = (box.select_one("#block_price")
                    or box.select_one("span.price")
                    or box.select_one("span[id*='price']:not(#old_price)")
                    or box.select_one("span"))
            if cand:
                price_raw = parse_price_digits(cand.get_text(" ", strip=True))
                if price_raw:
                    price_src = "prod_price_box"

    # 3) meta/itemprop (на случай микроразметки)
    if not price_raw:
        mp = main.select_one('[itemprop="price"][content]') or main.find("meta", attrs={"itemprop": "price"})
        if mp:
            price_raw = parse_price_digits(mp.get("content") or mp.get_text(" ", strip=True))
            if price_raw:
                price_src = "itemprop_price"

    if not price_raw:
        meta = main.find("meta", attrs={"property": re.compile(r"^(?:product|og):price:amount$", re.I)})
        if meta and meta.get("content"):
            price_raw = parse_price_digits(meta["content"])
            if price_raw:
                price_src = "meta_price_amount"

    # 4) Последний fallback: ищем "NNN тг" ТОЛЬКО внутри productfull
    if not price_raw:
        price_raw = parse_price_tenge(main.get_text(" ", strip=True))
        if price_raw:
            price_src = "text_tenge"

    if VERBOSE and price_src:
        log(f"[price] src={price_src} raw={price_raw} url={url}")

    # Availability
    available = True
    low = s.get_text(" ", strip=True).lower()
    if "нет в наличии" in low or "отсутств" in low:
        available = False

    return {
        "sku": sku.strip(),
        "title": title,
        "desc": desc_txt.strip(),
        "pics": pics,
        "pic": pic,
        "pics": pics,
        "params": [(k, v) for (k, v) in params2 if not re.fullmatch(r"\d{1,4}", k.strip())],
        "url": url,
        "price_raw": int(price_raw or 0),
        "available": bool(available),
    }


def discover_relevant_category_urls() -> List[str]:
    # Берём ссылки из /goods.html и главной, фильтруем по словам в тексте ссылки или в URL.
    seeds = [f"{BASE_URL}/", f"{BASE_URL}/goods.html"]
    pages: List[Tuple[str, BeautifulSoup]] = []
    for u in seeds:
        b = http_get(u, tries=3)
        if b:
            pages.append((u, soup_of(b)))
    if not pages:
        return []

    kws = [k.strip() for k in COPYLINE_INCLUDE_PREFIXES if k.strip()]
    urls: List[str] = []
    seen = set()

    for base, s in pages:
        for a in s.find_all("a", href=True):
            txt = safe_str(a.get_text(" ", strip=True) or "")
            absu = requests.compat.urljoin(base, safe_str(a["href"]))
            if "copyline.kz" not in absu:
                continue
            if "/goods/" not in absu and not absu.endswith("/goods.html"):
                continue

            ok = False
            for kw in kws:
                if re.search(r"(?i)(?<!\w)" + re.escape(kw).replace(r"\ ", " ") + r"(?!\w)", txt):
                    ok = True
                    break

            if not ok:
                slug = absu.lower()
                if any(h in slug for h in [
                    "drum", "developer", "fuser", "toner", "cartridge",
                    "драм", "девелопер", "фьюзер", "термоблок", "термоэлемент", "cartridg",
                    "кабель", "cable",
                ]):
                    ok = True

            if ok and absu not in seen:
                seen.add(absu)
                urls.append(absu)

    return list(dict.fromkeys(urls))


def _category_next_url(s: BeautifulSoup, page_url: str) -> Optional[str]:
    ln = s.find("link", attrs={"rel": "next"})
    if ln and ln.get("href"):
        return requests.compat.urljoin(page_url, safe_str(ln["href"]))
    a = s.find("a", class_=lambda c: c and "next" in safe_str(c).lower())
    if a and a.get("href"):
        return requests.compat.urljoin(page_url, safe_str(a["href"]))
    for a in s.find_all("a", href=True):
        txt = safe_str(a.get_text(" ", strip=True) or "").lower()
        if txt in ("следующая", "вперед", "вперёд", "next", ">"):
            return requests.compat.urljoin(page_url, safe_str(a["href"]))
    return None


def collect_product_urls(category_url: str, limit_pages: int) -> List[str]:
    # Собирает ссылки на товары внутри категории, проходя пагинацию.
    urls: List[str] = []
    seen_pages = set()
    page = category_url
    pages_done = 0

    while page and pages_done < limit_pages:
        if page in seen_pages:
            break
        seen_pages.add(page)

        _sleep_jitter(REQUEST_DELAY_MS)
        b = http_get(page, tries=3)
        if not b:
            break
        s = soup_of(b)

        for a in s.find_all("a", href=True):
            absu = requests.compat.urljoin(page, safe_str(a["href"]))
            if PRODUCT_RE.search(absu):
                urls.append(absu)

        page = _category_next_url(s, page)
        pages_done += 1

    return list(dict.fromkeys(urls))


def build_site_index(want_keys: Optional[Set[str]] = None) -> tuple[Dict[str, Dict[str, Any]], Dict[str, Dict[str, Any]]]:
    """CopyLine site index.
    Вместо парсинга категорий используем поиск по артикулу:
    https://copyline.kz/product-search/result.html (POST, поле search).
    Так стабильнее и совпадает с vendorCode_raw из XLSX."""
    if NO_CRAWL:
        log("[site] NO_CRAWL=1 -> skip site parsing")
        return {}, {}
    if not want_keys:
        log("[site] want_keys empty -> skip site parsing")
        return {}, {}

    deadline = datetime.utcnow() + timedelta(minutes=MAX_CRAWL_MINUTES)
    sku_index: Dict[str, Dict[str, Any]] = {}
    title_index: Dict[str, Dict[str, Any]] = {}

    def variants_for(raw: str) -> Set[str]:
        raw = safe_str(raw).strip()
        out: Set[str] = set()
        if not raw:
            return out
        out.add(norm_ascii(raw))
        out.add(norm_ascii(raw.replace("-", "")))
        z0 = raw.lstrip("0")
        if z0 and z0 != raw:
            out.add(norm_ascii(z0))
            out.add(norm_ascii(z0.replace("-", "")))
        if re.fullmatch(r"[Cc]\d+", raw):
            out.add(norm_ascii(raw[1:]))
        if re.fullmatch(r"\d+", raw):
            out.add(norm_ascii("C" + raw))
        return out

    from concurrent.futures import ThreadPoolExecutor, as_completed
    want_list = list(dict.fromkeys([safe_str(x).strip() for x in want_keys if safe_str(x).strip()]))

    log(f"[site] search-mode keys={len(want_list)} workers={MAX_WORKERS} minutes={MAX_CRAWL_MINUTES}")

    def job(sku: str) -> Optional[Dict[str, Any]]:
        if datetime.utcnow() > deadline:
            return None
        try:
            return _search_copyline_for_sku(sku)
        except Exception:
            return None

    with ThreadPoolExecutor(max_workers=max(1, MAX_WORKERS)) as ex:
        futures = {ex.submit(job, sku): sku for sku in want_list}
        for fut in as_completed(futures):
            if datetime.utcnow() > deadline:
                break
            sku = futures[fut]
            out = None
            try:
                out = fut.result()
            except Exception:
                out = None
            if not out:
                continue

            best_pic = _pick_best_copyline_pic(list(out.get("pics") or []) + ([safe_str(out.get("pic"))] if out.get("pic") else []))
            if best_pic:
                out["pic"] = best_pic
                out["pics"] = [best_pic]
            else:
                out["pic"] = ""
                out["pics"] = []

            for k in variants_for(sku):
                if k and k not in sku_index:
                    sku_index[k] = out

            t = safe_str(out.get("title") or "")
            if t:
                tk = norm_ascii(title_clean(t))
                if tk and tk not in title_index:
                    title_index[tk] = out
                tk30 = norm_ascii(title_clean(t[:30]))
                if tk30 and tk30 not in title_index:
                    title_index[tk30] = out

    log(f"[site] indexed={len(sku_index)} title_index={len(title_index)}")
    return sku_index, title_index

def next_run_dom_1_10_20_at_hour(now_local: datetime, hour: int) -> datetime:
    # now_local — наивный datetime в Алматы
    y = now_local.year
    m = now_local.month

    def candidates_for_month(yy: int, mm: int) -> List[datetime]:
        return [datetime(yy, mm, d, hour, 0, 0) for d in (1, 10, 20)]

    cands = [dt for dt in candidates_for_month(y, m) if dt > now_local]
    if cands:
        return min(cands)

    # следующий месяц
    if m == 12:
        y2, m2 = y + 1, 1
    else:
        y2, m2 = y, m + 1
    return min(candidates_for_month(y2, m2))


# -----------------------------
# Main
# -----------------------------



def parse_sitemap_xml_urls(xml_bytes: bytes) -> List[Dict[str, str]]:
    """Fallback: sitemap.xml (urlset). Названий нет — фильтр сделаем уже по title со страницы товара."""
    txt = (xml_bytes or b"").decode("utf-8", errors="ignore")
    urls = re.findall(r"<loc>\s*([^<\s]+)\s*</loc>", txt, flags=re.I)
    out: List[Dict[str, str]] = []
    seen = set()
    for u in urls:
        u = safe_str(u).strip()
        if not u:
            continue
        if "/goods/" not in u or not re.search(r"\.html(?:\?|$)", u, flags=re.I):
            continue
        if u in seen:
            continue
        seen.add(u)
        out.append({"url": u, "title": ""})
    return out

def _abs_url(href: str) -> str:
    href = safe_str(href).strip()
    if not href:
        return ""
    if href.startswith("http://") or href.startswith("https://"):
        return href
    if href.startswith("//"):
        return "https:" + href
    if href.startswith("/"):
        return BASE_URL + href
    return BASE_URL.rstrip("/") + "/" + href.lstrip("/")


def parse_sitemap_products(html_bytes: bytes) -> List[Dict[str, str]]:
    """Парсим html-sitemap и берём только товары (ссылки /goods/*.html),
    затем фильтруем по COPYLINE_INCLUDE_PREFIXES (с начала строки).
    Возвращаем список dict: {url,title}.
    """
    s = soup_of(html_bytes)
    out: List[Dict[str, str]] = []
    seen = set()

    for a in s.find_all("a"):
        href = safe_str(a.get("href"))
        title = title_clean(safe_str(a.get_text(" ", strip=True)))
        if not href or not title:
            continue

        url = _abs_url(href)
        if "/goods/" not in url or not re.search(r"\.html(?:\?|$)", url, flags=re.I):
            continue
        if not _is_allowed_prefix(title):
            continue

        key = url
        if key in seen:
            continue
        seen.add(key)
        out.append({"url": url, "title": title})

    return out

def main() -> int:
    build_time = now_almaty()
    next_run = next_run_dom_1_10_20_at_hour(build_time, 3)

    if NO_CRAWL:
        raise RuntimeError("NO_CRAWL=1: crawl отключён, а режим CopyLine теперь sitemap.")

    sitemap_bytes = http_get(SITEMAP_URL, tries=3, min_bytes=20_000)
    products: List[Dict[str, str]] = []
    if sitemap_bytes:
        products = parse_sitemap_products(sitemap_bytes)
    else:
        # fallback: sitemap.xml
        xml_bytes = http_get(f"{BASE_URL}/sitemap.xml", tries=3, min_bytes=5_000)
        if not xml_bytes:
            raise RuntimeError("CopyLine: не удалось скачать sitemap (HTML) и sitemap.xml.")
        products = parse_sitemap_xml_urls(xml_bytes)
    before = len(products)

    # Собираем offers (важно: стабильные oid!)
    out_offers: List[OfferOut] = []
    seen_oids = set()

    deadline = datetime.utcnow() + timedelta(minutes=MAX_CRAWL_MINUTES)
    from concurrent.futures import ThreadPoolExecutor, as_completed

    def _mk_oid(sku: str) -> str:
        sku = safe_str(sku).strip()
        sku = re.sub(r"[^A-Za-z0-9\-\._/]", "", sku)
        return "CL" + sku

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futs = {ex.submit(parse_product_page, p["url"]): p for p in products}
        for fut in as_completed(futs):
            if datetime.utcnow() > deadline:
                log("[site] deadline reached -> stop parsing products")
                break

            p = futs[fut]
            got = None
            try:
                got = fut.result()
            except Exception as e:
                log(f"[site] parse fail: {p.get('url')} | {e}")
                continue

            if not got:
                continue

            name = safe_str(got.get("title") or p.get("title") or "").strip()
            if not _is_allowed_prefix(name):
                continue

            sku = safe_str(got.get("sku") or "").strip()
            if not sku:
                continue

            oid = _mk_oid(sku)
            if oid in seen_oids:
                continue
            seen_oids.add(oid)

            native_desc = safe_str(got.get("desc") or "").strip() or name

            # Params
            params: List[Tuple[str, str]] = []
            for (k, v) in (got.get("params") or []):
                kk = safe_str(k).strip()
                vv = safe_str(v).strip()
                if kk and vv:
                    params.append((kk, vv))

            # Минимальные характеристики (чтобы было полезно даже если на странице пусто)
            kind = _derive_kind(name)
            if kind:
                params = _merge_params(params, [("Тип", kind)])

            raw_price = int(got.get("price_raw") or 0)
            price = compute_price(raw_price)

            pictures = got.get("pics") or []
            if not isinstance(pictures, list):
                pictures = [safe_str(pictures)]
            pic = safe_str(got.get("pic") or "").strip()
            if (not pictures) and pic:
                pictures = [pic]
            pictures = _pick_copyline_best_picture([safe_str(p).strip() for p in pictures if safe_str(p).strip()])

            out_offers.append(
                OfferOut(
                    oid=oid,
                    available=True,
                    name=name,
                    price=price,
                    pictures=_copyline_full_only(pictures),
                    vendor="",  # бренд будет выбран ядром; если не найдётся — упадём на PUBLIC_VENDOR
                    params=params,
                    native_desc=native_desc,
                )
            )

    # стабильный порядок
    out_offers.sort(key=lambda o: o.oid)

    after = len(out_offers)
    in_true = sum(1 for o in out_offers if o.available)
    in_false = after - in_true

    feed_meta = make_feed_meta(
        supplier=SUPPLIER_NAME,
        supplier_url=os.getenv("SUPPLIER_URL", SUPPLIER_URL_DEFAULT),
        build_time=build_time,
        next_run=next_run,
        before=before,
        after=after,
        in_true=in_true,
        in_false=in_false,
    )

    header = make_header(build_time, encoding=OUTPUT_ENCODING)
    footer = make_footer()

    offers_xml = "\n\n".join(
        [o.to_xml(currency_id=CURRENCY_ID_DEFAULT, public_vendor=PUBLIC_VENDOR) for o in out_offers]
    )

    full = header + "\n" + feed_meta + "\n\n" + offers_xml + "\n" + footer
    full = ensure_footer_spacing(full)
    validate_cs_yml(full)
    changed = write_if_changed(OUT_FILE, full, encoding=OUTPUT_ENCODING)

    log(
        f"[build_copyline] OK | offers_in={before} | offers_out={after} | in_true={in_true} | in_false={in_false} | "
        f"crawl={'no' if NO_CRAWL else 'yes'} | changed={'yes' if changed else 'no'} | file={OUT_FILE}",
        flush=True,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
